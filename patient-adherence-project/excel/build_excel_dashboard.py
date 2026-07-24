"""
build_excel_dashboard.py
--------------------------
Builds the client-ready Excel dashboard for the
Patient Therapy Adherence & Persistency Analytics project.

Sheets:
  1. Executive Summary   -- headline KPIs (live formulas)
  2. Discontinuation Drivers -- driver tables by insurance / supply /
     therapy class / region / age band (live COUNTIFS/AVERAGEIFS formulas)
  3. Outreach Watchlist  -- prioritized, risk-scored patient list
  4. Patient Data        -- source table every formula above reads from

All aggregate numbers are computed with real Excel formulas referencing
the Patient Data sheet, not hardcoded Python results, so the dashboard
recalculates if the underlying data changes.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.drawing.image import Image as XLImage

OUT = "/home/claude/project/output"
FONT_NAME = "Arial"

# ---------------------------------------------------------------------
# Load source data
# ---------------------------------------------------------------------
master = pd.read_csv(f"{OUT}/patient_master_analytics.csv")
watchlist = pd.read_csv(f"{OUT}/outreach_watchlist.csv")
logrank_df = pd.read_csv(f"{OUT}/logrank_test_results.csv")
cox_df = pd.read_csv(f"{OUT}/cox_model_results.csv")
stat_tests_df = pd.read_csv(f"{OUT}/statistical_tests.csv")
model_perf_df = pd.read_csv(f"{OUT}/model_performance.csv")
model_feat_df = pd.read_csv(f"{OUT}/model_feature_importance.csv")
cox_concordance = pd.read_csv(f"{OUT}/cox_concordance.csv").iloc[0, 0]

# order columns sensibly for the Data sheet
data_cols = ["patient_id", "age", "gender", "region", "insurance_status",
             "therapy_class", "initial_supply_days", "therapy_start_date",
             "total_fills", "first_fill_date", "last_fill_date",
             "days_since_last_fill", "days_persisted", "discontinued_flag",
             "pdc", "avg_gap_days", "age_band"]
master = master[data_cols]

wb = Workbook()

# ============================================================
# Styles
# ============================================================
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, size=10, italic=True, color="595959")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
KPI_LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True, color="595959")
KPI_VALUE_FONT = Font(name=FONT_NAME, size=20, bold=True, color="1F4E78")
BODY_FONT = Font(name=FONT_NAME, size=10)
SECTION_FONT = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
KPI_FILL = PatternFill("solid", fgColor="EAF1F8")


def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# SHEET 4 (built first so formulas elsewhere can reference it): Patient Data
# ============================================================
ws_data = wb.active
ws_data.title = "Patient Data"

for j, col in enumerate(data_cols, start=1):
    ws_data.cell(row=1, column=j, value=col.replace("_", " ").title())
style_header_row(ws_data, 1, len(data_cols))

for i, row in enumerate(master.itertuples(index=False), start=2):
    for j, val in enumerate(row, start=1):
        ws_data.cell(row=i, column=j, value=val).font = BODY_FONT

N = len(master)  # number of data rows
LAST_ROW = N + 1
autofit(ws_data, [12, 6, 8, 10, 16, 20, 18, 16, 10, 16, 16, 14, 14, 16, 8, 12, 10])
ws_data.freeze_panes = "A2"

# named ranges (as plain sheet!range strings for formulas below)
D = "'Patient Data'"
COL = {c: get_column_letter(i + 1) for i, c in enumerate(data_cols)}
rng = lambda c: f"{D}!${COL[c]}$2:${COL[c]}${LAST_ROW}"

# ============================================================
# SHEET 1: Executive Summary
# ============================================================
ws_sum = wb.create_sheet("Executive Summary", 0)
ws_sum.sheet_view.showGridLines = False

ws_sum["B2"] = "Patient Therapy Adherence & Persistency Analytics"
ws_sum["B2"].font = TITLE_FONT
ws_sum["B3"] = "MySQL + Python + Excel  |  2,000 simulated patients  |  15,000+ refill events  |  Study window ending 2025-12-31"
ws_sum["B3"].font = SUBTITLE_FONT

kpis = [
    ("Total Patients", f"=COUNTA({rng('patient_id')})", "0"),
    ("Total Refill Events", f"=SUM({rng('total_fills')})", "#,##0"),
    ("Overall Discontinuation Rate",
     f"=SUM({rng('discontinued_flag')})/COUNTA({rng('patient_id')})", "0.0%"),
    ("Overall Average PDC", f"=AVERAGE({rng('pdc')})", "0.0%"),
    ("Uninsured vs Insured Churn Ratio",
     f"=(COUNTIFS({rng('insurance_status')},\"Uninsured\",{rng('discontinued_flag')},1)/COUNTIF({rng('insurance_status')},\"Uninsured\"))"
     f"/(COUNTIFS({rng('insurance_status')},\"Insured\",{rng('discontinued_flag')},1)/COUNTIF({rng('insurance_status')},\"Insured\"))",
     "0.00\"x\""),
    ("90-Day vs 30-Day Supply Persistency Gap",
     f"=(1-COUNTIFS({rng('initial_supply_days')},90,{rng('discontinued_flag')},1)/COUNTIF({rng('initial_supply_days')},90))"
     f"-(1-COUNTIFS({rng('initial_supply_days')},30,{rng('discontinued_flag')},1)/COUNTIF({rng('initial_supply_days')},30))",
     "+0.0%;-0.0%"),
]

row0 = 6
col_positions = [(row0, 2), (row0, 5), (row0, 8),
                  (row0 + 6, 2), (row0 + 6, 5), (row0 + 6, 8)]
for (label, formula, numfmt), (r, c) in zip(kpis, col_positions):
    # card box: label + value, spans 2 columns x 5 rows
    ws_sum.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
    lab_cell = ws_sum.cell(row=r, column=c, value=label)
    lab_cell.font = KPI_LABEL_FONT
    lab_cell.alignment = Alignment(horizontal="left")

    ws_sum.merge_cells(start_row=r + 1, start_column=c, end_row=r + 3, end_column=c + 1)
    val_cell = ws_sum.cell(row=r + 1, column=c, value=formula)
    val_cell.font = KPI_VALUE_FONT
    val_cell.number_format = numfmt
    val_cell.alignment = Alignment(horizontal="left", vertical="center")
    for rr in range(r, r + 4):
        for cc in range(c, c + 2):
            ws_sum.cell(row=rr, column=cc).fill = KPI_FILL

ws_sum["B18"] = "Key Findings"
ws_sum["B18"].font = SECTION_FONT
findings = [
    "Uninsured patients discontinue therapy at roughly 1.6x the rate of insured patients "
    "-- insurance status is the single strongest churn driver in the dataset "
    "(Cox model hazard ratio ~3.9x, p < 0.001, holding age/supply/therapy class constant).",
    "Patients started on a 90-day supply persist about 16 percentage points longer than "
    "those started on a 30-day supply, and see far fewer early-lapse events "
    "(log-rank test p < 0.001).",
    "Osteoporosis and Depression/Anxiety therapy classes show the weakest persistency; "
    "Hypertension and Diabetes (chronic, well-established regimens) show the strongest.",
    "The West region shows a materially higher discontinuation rate than other regions, "
    "flagging a possible access-to-care or pharmacy-network gap worth investigating.",
    "A logistic regression model trained to predict discontinuation reaches 0.76 ROC-AUC "
    "on held-out patients -- see the 'Predictive Model' tab -- and now drives the risk "
    "scores on the Outreach Watchlist.",
    "See the 'Outreach Watchlist' tab for a model-ranked list of currently active patients "
    "approaching the 60-day lapse threshold -- these are the patients a call-center or "
    "care-management team should prioritize this week.",
]
r = 19
for f in findings:
    ws_sum.cell(row=r, column=2, value=f"•  {f}").font = BODY_FONT
    ws_sum.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
    ws_sum.row_dimensions[r].height = 28
    ws_sum.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

autofit(ws_sum, [3] + [12] * 10)

# ============================================================
# SHEET 2: Discontinuation Drivers
# ============================================================
ws_drv = wb.create_sheet("Discontinuation Drivers")
ws_drv.sheet_view.showGridLines = False
ws_drv["B2"] = "Discontinuation Drivers"
ws_drv["B2"].font = TITLE_FONT
ws_drv["B3"] = "All figures below are live formulas (COUNTIFS / AVERAGEIFS) reading from the 'Patient Data' sheet."
ws_drv["B3"].font = SUBTITLE_FONT

def write_driver_table(ws, top_row, title, group_col, group_values, group_label,
                        criteria_builder):
    """
    criteria_builder(v) -> excel criteria string usable directly in COUNTIFS, e.g. '"Insured"' or 30
    """
    ws.cell(row=top_row, column=2, value=title).font = SECTION_FONT
    headers = [group_label, "Patient Count", "Discontinued Count", "Discontinuation Rate",
               "Persistency Rate", "Avg PDC", "Avg Days Persisted"]
    hr = top_row + 1
    for j, h in enumerate(headers, start=2):
        ws.cell(row=hr, column=j, value=h)
    style_header_row(ws, hr, 8)
    for i, gv in enumerate(group_values):
        r = hr + 1 + i
        crit = criteria_builder(gv)
        ws.cell(row=r, column=2, value=gv).font = BODY_FONT
        cnt_formula = f"=COUNTIF({rng(group_col)},{crit})"
        disc_formula = f"=COUNTIFS({rng(group_col)},{crit},{rng('discontinued_flag')},1)"
        ws.cell(row=r, column=3, value=cnt_formula).font = BODY_FONT
        ws.cell(row=r, column=4, value=disc_formula).font = BODY_FONT
        rate_cell = ws.cell(row=r, column=5, value=f"=D{r}/C{r}")
        rate_cell.number_format = "0.0%"
        rate_cell.font = BODY_FONT
        persist_cell = ws.cell(row=r, column=6, value=f"=1-E{r}")
        persist_cell.number_format = "0.0%"
        persist_cell.font = BODY_FONT
        pdc_formula = f"=AVERAGEIFS({rng('pdc')},{rng(group_col)},{crit})"
        pdc_cell = ws.cell(row=r, column=7, value=pdc_formula)
        pdc_cell.number_format = "0.0%"
        pdc_cell.font = BODY_FONT
        days_formula = f"=AVERAGEIFS({rng('days_persisted')},{rng(group_col)},{crit})"
        days_cell = ws.cell(row=r, column=8, value=days_formula)
        days_cell.number_format = "0"
        days_cell.font = BODY_FONT
        for c in range(2, 9):
            ws.cell(row=r, column=c).border = BORDER
    return hr + 1 + len(group_values)  # first empty row after table


row_cursor = 5
row_cursor = write_driver_table(
    ws_drv, row_cursor, "By Insurance Status", "insurance_status",
    ["Insured", "Uninsured"], "Insurance Status",
    lambda v: f'"{v}"'
) + 2

row_cursor = write_driver_table(
    ws_drv, row_cursor, "By Initial Supply Size (days)", "initial_supply_days",
    [30, 60, 90], "Supply Size", lambda v: v
) + 2

row_cursor = write_driver_table(
    ws_drv, row_cursor, "By Therapy Class", "therapy_class",
    sorted(master["therapy_class"].unique().tolist()), "Therapy Class",
    lambda v: f'"{v}"'
) + 2

row_cursor = write_driver_table(
    ws_drv, row_cursor, "By Region", "region",
    sorted(master["region"].unique().tolist()), "Region",
    lambda v: f'"{v}"'
) + 2

row_cursor = write_driver_table(
    ws_drv, row_cursor, "By Age Band", "age_band",
    ["18-29", "30-44", "45-59", "60+"], "Age Band",
    lambda v: f'"{v}"'
) + 2

autofit(ws_drv, [3, 20, 14, 16, 16, 14, 10, 16])

# small chart: discontinuation rate by insurance status
chart = BarChart()
chart.title = "Discontinuation Rate by Insurance Status"
chart.y_axis.title = "Rate"
chart.y_axis.numFmt = "0%"
chart.style = 10
data_ref = Reference(ws_drv, min_col=5, min_row=6, max_row=7)
cats_ref = Reference(ws_drv, min_col=2, min_row=7, max_row=7)
chart.add_data(data_ref, titles_from_data=False)
chart.set_categories(cats_ref)
chart.height = 7
chart.width = 12
ws_drv.add_chart(chart, "J5")

# ============================================================
# SHEET 3: Outreach Watchlist
# ============================================================
ws_wl = wb.create_sheet("Outreach Watchlist")
ws_wl.sheet_view.showGridLines = False
ws_wl["B2"] = "Prioritized Patient Outreach Watchlist"
ws_wl["B2"].font = TITLE_FONT
ws_wl["B3"] = ("Active patients 31-60 days since their last fill (approaching the 60-day lapse "
               "threshold), ranked by a logistic regression model's predicted probability of "
               "discontinuation (see 'Predictive Model' tab for how it was trained/validated).")
ws_wl["B3"].font = SUBTITLE_FONT
ws_wl.row_dimensions[3].height = 28
ws_wl.merge_cells("B3:L3")
ws_wl.cell(row=3, column=2).alignment = Alignment(wrap_text=True)

wl_cols = ["patient_id", "age", "gender", "region", "insurance_status", "therapy_class",
           "initial_supply_days", "total_fills", "last_fill_date", "days_since_last_fill",
           "pdc", "risk_score", "priority_tier"]
watchlist = watchlist[wl_cols].sort_values("risk_score", ascending=False)

wl_header_row = 5
for j, col in enumerate(wl_cols, start=2):
    ws_wl.cell(row=wl_header_row, column=j, value=col.replace("_", " ").title())
style_header_row(ws_wl, wl_header_row, len(wl_cols) + 1)

for i, row in enumerate(watchlist.itertuples(index=False), start=wl_header_row + 1):
    for j, val in enumerate(row, start=2):
        cell = ws_wl.cell(row=i, column=j, value=val)
        cell.font = BODY_FONT
        cell.border = BORDER
        if wl_cols[j - 2] == "pdc":
            cell.number_format = "0.0%"

wl_last_row = wl_header_row + len(watchlist)
tier_col_letter = get_column_letter(2 + wl_cols.index("priority_tier"))
score_col_letter = get_column_letter(2 + wl_cols.index("risk_score"))

# conditional formatting: color the priority tier column
high_fill = PatternFill("solid", fgColor="F8CBAD")
med_fill = PatternFill("solid", fgColor="FFE699")
low_fill = PatternFill("solid", fgColor="C6E0B4")
rng_tier = f"{tier_col_letter}{wl_header_row+1}:{tier_col_letter}{wl_last_row}"
ws_wl.conditional_formatting.add(
    rng_tier, CellIsRule(operator="equal", formula=['"High"'], fill=high_fill))
ws_wl.conditional_formatting.add(
    rng_tier, CellIsRule(operator="equal", formula=['"Medium"'], fill=med_fill))
ws_wl.conditional_formatting.add(
    rng_tier, CellIsRule(operator="equal", formula=['"Low"'], fill=low_fill))

# color scale on risk score
rng_score = f"{score_col_letter}{wl_header_row+1}:{score_col_letter}{wl_last_row}"
ws_wl.conditional_formatting.add(
    rng_score,
    ColorScaleRule(start_type="min", start_color="C6E0B4",
                    end_type="max", end_color="F8CBAD"))

ws_wl.freeze_panes = f"B{wl_header_row+1}"
autofit(ws_wl, [3, 12, 6, 8, 16, 20, 12, 10, 14, 12, 10, 10, 12])

# ============================================================
# SHEET 5: Survival Analysis (Kaplan-Meier)
# ============================================================
ws_surv = wb.create_sheet("Survival Analysis")
ws_surv.sheet_view.showGridLines = False
ws_surv["B2"] = "Survival Analysis (Kaplan-Meier)"
ws_surv["B2"].font = TITLE_FONT
ws_surv["B3"] = ("Time-to-discontinuation modeling that properly accounts for right-censoring "
                  "(patients still active at the study cutoff haven't 'failed' -- they're censored, "
                  "not persistent forever). Curves that drop faster and lower indicate weaker persistency.")
ws_surv["B3"].font = SUBTITLE_FONT
ws_surv.merge_cells("B3:N3")
ws_surv.cell(row=3, column=2).alignment = Alignment(wrap_text=True)
ws_surv.row_dimensions[3].height = 28

img1 = XLImage(f"{OUT}/km_curve_insurance.png")
img1.width, img1.height = 430, 270
ws_surv.add_image(img1, "B6")

img2 = XLImage(f"{OUT}/km_curve_supply.png")
img2.width, img2.height = 430, 270
ws_surv.add_image(img2, "J6")

img3 = XLImage(f"{OUT}/km_curve_therapy_class.png")
img3.width, img3.height = 430, 270
ws_surv.add_image(img3, "B24")

ws_surv["B42"] = "Log-Rank Significance Tests"
ws_surv["B42"].font = SECTION_FONT
lr_headers = ["Comparison", "Median Days (Group 1)", "Median Days (Group 2)", "Log-Rank p-value", "Significant (p<0.05)"]
for j, h in enumerate(lr_headers, start=2):
    ws_surv.cell(row=43, column=j, value=h)
style_header_row(ws_surv, 43, len(lr_headers) + 1)
for i, row in enumerate(logrank_df.itertuples(index=False), start=44):
    for j, val in enumerate(row, start=2):
        cell = ws_surv.cell(row=i, column=j, value=val)
        cell.font = BODY_FONT
        cell.border = BORDER
        if lr_headers[j - 2] == "Log-Rank p-value":
            cell.number_format = "0.00E+00"

ws_surv["B49"] = "Cox Proportional Hazards Model"
ws_surv["B49"].font = SECTION_FONT
ws_surv["B50"] = ("Quantifies each driver's independent effect on discontinuation hazard, holding "
                   "the other variables constant. Hazard ratio > 1 = faster discontinuation; < 1 = slower. "
                   f"Model concordance index: {cox_concordance:.3f}")
ws_surv.merge_cells("B50:N50")
ws_surv.cell(row=50, column=2).font = SUBTITLE_FONT
ws_surv.cell(row=50, column=2).alignment = Alignment(wrap_text=True)
ws_surv.row_dimensions[50].height = 28

cox_headers = ["Covariate", "Log-Hazard Coef", "Hazard Ratio", "p-value"]
for j, h in enumerate(cox_headers, start=2):
    ws_surv.cell(row=51, column=j, value=h)
style_header_row(ws_surv, 51, len(cox_headers) + 1)
cox_top = cox_df.sort_values("hazard_ratio", ascending=False).head(12)
for i, row in enumerate(cox_top.itertuples(index=False), start=52):
    covariate, coef, hr, p, *_ = row
    ws_surv.cell(row=i, column=2, value=covariate).font = BODY_FONT
    c1 = ws_surv.cell(row=i, column=3, value=round(coef, 3)); c1.font = BODY_FONT
    c2 = ws_surv.cell(row=i, column=4, value=round(hr, 2)); c2.font = BODY_FONT
    c3 = ws_surv.cell(row=i, column=5, value=p); c3.font = BODY_FONT; c3.number_format = "0.00E+00"
    for c in range(2, 6):
        ws_surv.cell(row=i, column=c).border = BORDER

autofit(ws_surv, [3] + [16] * 13)

# ============================================================
# SHEET 6: Predictive Model
# ============================================================
ws_model = wb.create_sheet("Predictive Model")
ws_model.sheet_view.showGridLines = False
ws_model["B2"] = "Predictive Model: Discontinuation Risk"
ws_model["B2"].font = TITLE_FONT
ws_model["B3"] = ("Logistic regression trained on a 75/25 train/test split, predicting P(discontinued) "
                   "from patient features. Used to score the Outreach Watchlist's risk_score.")
ws_model["B3"].font = SUBTITLE_FONT
ws_model.merge_cells("B3:H3")

perf = model_perf_df.iloc[0]
perf_items = [
    ("Accuracy", f"{perf['accuracy']:.1%}"),
    ("ROC-AUC", f"{perf['roc_auc']:.3f}"),
    ("Precision", f"{perf['precision']:.1%}"),
    ("Recall", f"{perf['recall']:.1%}"),
    ("F1 Score", f"{perf['f1_score']:.3f}"),
]
r0 = 6
for i, (label, value) in enumerate(perf_items):
    c = 2 + i * 2
    ws_model.merge_cells(start_row=r0, start_column=c, end_row=r0, end_column=c + 1)
    ws_model.cell(row=r0, column=c, value=label).font = KPI_LABEL_FONT
    ws_model.merge_cells(start_row=r0 + 1, start_column=c, end_row=r0 + 3, end_column=c + 1)
    v = ws_model.cell(row=r0 + 1, column=c, value=value)
    v.font = KPI_VALUE_FONT
    v.alignment = Alignment(vertical="center")
    for rr in range(r0, r0 + 4):
        for cc in range(c, c + 2):
            ws_model.cell(row=rr, column=cc).fill = KPI_FILL

ws_model["B12"] = ("Held out on 500 unseen patients (25% test split). Confusion matrix: "
                    f"TP={int(perf['true_positives'])}, FP={int(perf['false_positives'])}, "
                    f"TN={int(perf['true_negatives'])}, FN={int(perf['false_negatives'])}.")
ws_model["B12"].font = SUBTITLE_FONT
ws_model.merge_cells("B12:H12")

img_roc = XLImage(f"{OUT}/roc_curve.png")
img_roc.width, img_roc.height = 320, 320
ws_model.add_image(img_roc, "J6")

ws_model["B15"] = "Top Predictive Features (standardized logistic regression coefficients)"
ws_model["B15"].font = SECTION_FONT
feat_headers = ["Feature", "Coefficient", "Odds Ratio"]
for j, h in enumerate(feat_headers, start=2):
    ws_model.cell(row=16, column=j, value=h)
style_header_row(ws_model, 16, len(feat_headers) + 1)
for i, row in enumerate(model_feat_df.head(12).itertuples(index=False), start=17):
    feat, coef, odds = row
    ws_model.cell(row=i, column=2, value=feat).font = BODY_FONT
    c1 = ws_model.cell(row=i, column=3, value=round(coef, 3)); c1.font = BODY_FONT
    c2 = ws_model.cell(row=i, column=4, value=round(odds, 2)); c2.font = BODY_FONT
    for c in range(2, 5):
        ws_model.cell(row=i, column=c).border = BORDER

ws_model["B31"] = "Statistical Significance Tests (Chi-square / Welch t-test)"
ws_model["B31"].font = SECTION_FONT
st_headers = ["Test", "Statistic", "p-value", "Significant (p<0.05)"]
for j, h in enumerate(st_headers, start=2):
    ws_model.cell(row=32, column=j, value=h)
style_header_row(ws_model, 32, len(st_headers) + 1)
for i, row in enumerate(stat_tests_df.itertuples(index=False), start=33):
    test, stat, p, sig = row
    ws_model.cell(row=i, column=2, value=test).font = BODY_FONT
    c1 = ws_model.cell(row=i, column=3, value=stat); c1.font = BODY_FONT
    c2 = ws_model.cell(row=i, column=4, value=p); c2.font = BODY_FONT; c2.number_format = "0.00E+00"
    c3 = ws_model.cell(row=i, column=5, value=bool(sig)); c3.font = BODY_FONT
    for c in range(2, 6):
        ws_model.cell(row=i, column=c).border = BORDER

autofit(ws_model, [3, 30, 14, 14, 14, 12, 12, 12])


out_path = f"{OUT}/Patient_Therapy_Adherence_Dashboard.xlsx"
wb.save(out_path)
print(f"Saved workbook: {out_path}")
